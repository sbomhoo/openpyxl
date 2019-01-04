from openpyxl import Workbook       # 라이브러리 설치 pip install openpyxl
from openpyxl import load_workbook  # 엑셀읽어오기 위한 라이브러리



'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''
openpyxl 라이브러리를 이용하여 엑셀 파일을 읽어서 
초록의 각 셀의 내용들을 각 text file에 저장
'''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''''


def create_txt(row_index,row_contents):
    f=open("result/text{0}.txt".format(row_index),'w', encoding='UTF8')
    f.write(row_contents)
    
    
#Data파일 가져오기
wb = load_workbook('data_560.xlsx') 
sheet1 = wb.get_sheet_by_name("Sheet1") #읽을 Data파일 스프레드 시트 설정

sheet1.delete_rows(1)  #첫 행 헤더 부분 지우기


#새로운 엑셀 파일에 저장하기 위해 객체 생성
result = Workbook() 
result_sheet = result.active 


#읽어오기
for r in sheet1.rows:
    row_index = r[0].row   # r[컬럼번호].row 첫번째 컬럼의 행 번호를 저장
    row_contents = r[1].value #초록    r[1].value 두번쨰 칼럼의 셀 내용 저장
    #row_id = r[0].value  #논문 i     
    #row_yyyymm=r[2].value #발행연도   
    
    
    create_txt(row_index,row_contents)
    
    print(row_index)
    

wb.close()